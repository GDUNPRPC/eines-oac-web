#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AUDITOR LLIBRE DE VISITES - NUCLI WEB (Pyodide)
Mossos d'Esquadra - Regió Policial Central

Lògica d'auditoria EXTRETA de auditor.py (v5.5), SENSE Tkinter ni PIL, perquè
pugui executar-se dins del navegador amb Pyodide (WebAssembly).

La lògica de processament i validació és IDÈNTICA a la versió d'escriptori ja
validada. Només canvia la capa d'entrada/sortida:
  - Els dos Excels d'entrada (Visites, Entitats) arriben com a `bytes`.
  - Les sortides (Registre validat + informes) es retornen empaquetades en un
    ZIP (`bytes`), perquè poden ser diversos fitxers.
  - La configuració s'injecta amb set_config() (no es llegeix de config.json).

Cap dada surt del navegador: tot el processament passa a la màquina de l'usuari.
"""

import re
import unicodedata
import datetime
import zipfile
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================================
# CONFIGURACIÓ (injectada des del navegador)
# ============================================================================

FILAS_SKIP = 5
TIEMPO_ESPERA_MINIMO = 30
MINUTOS_LIMITE_TRAMITE = 15
MAPEO_COMISARIAS = {}
MAPEO_COLUMNES_VISITES = {}
MAPEO_COLUMNES_ENTITATS = {}
codigos_validacion = set()
LOGO_PATH = None  # ruta dins del FS virtual de Pyodide (opcional)


def set_config(cfg):
    """Injecta la configuració (secció 'auditor' del config.json)."""
    global FILAS_SKIP, TIEMPO_ESPERA_MINIMO, MINUTOS_LIMITE_TRAMITE
    global MAPEO_COMISARIAS, MAPEO_COLUMNES_VISITES, MAPEO_COLUMNES_ENTITATS
    global codigos_validacion
    cfg = cfg or {}
    FILAS_SKIP = cfg.get("filas_skip", 5)
    TIEMPO_ESPERA_MINIMO = cfg.get("tiempo_espera_minimo", 30)
    MINUTOS_LIMITE_TRAMITE = cfg.get("minutos_limite_tramite", 15)
    MAPEO_COMISARIAS = cfg.get("mapeo_comisarias", {}) or {}
    MAPEO_COLUMNES_VISITES = cfg.get("mapeo_columnes_visites", {}) or {}
    MAPEO_COLUMNES_ENTITATS = cfg.get("mapeo_columnes_entitats", {}) or {}
    codigos_validacion = set(cfg.get("codigos_validacion", []) or [])


def set_logo_path(path):
    global LOGO_PATH
    LOGO_PATH = path


# Mesos en català
MESES_CATALAN = {
    "January": "gener", "February": "febrer", "March": "març", "April": "abril",
    "May": "maig", "June": "juny", "July": "juliol", "August": "agost",
    "September": "setembre", "October": "octubre", "November": "novembre", "December": "desembre"
}
MESES_CATALAN_LONG = {
    "January": "de gener", "February": "de febrer", "March": "de març",
    "April": "d'abril", "May": "de maig", "June": "de juny",
    "July": "de juliol", "August": "d'agost", "September": "de setembre",
    "October": "d'octubre", "November": "de novembre", "December": "de desembre"
}

ESTIL_INFORMES = {
    "color_encabezado": "1F4E78",
    "color_texto_encabezado": "FFFFFF",
    "color_fila_par": "F2F2F2",
    "color_hora_modificar": "FFF2CC",
    "color_professional": "E2EFDA",
    "color_titulo": "2E5C8A",
    "fuente_titulo": 22,
    "fuente_subtitulo": 16,
    "fuente_encabezado": 13,
    "fuente_datos": 11,
    "altura_titulo": 30,
    "altura_encabezado": 25,
    "altura_datos": 20,
}


# ============================================================================
# UTILITATS
# ============================================================================

def _valor_celda_str(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str(val)


def _valor_any_tramite(tramite) -> str:
    for key in ("Unnamed: 2", "Any", "Año", "Year"):
        if key in tramite.index:
            val = tramite.get(key)
            if pd.notna(val) and str(val).strip():
                return _valor_celda_str(val)
    for col in tramite.index:
        if re.search(r"\b(any|año|year)\b", str(col), flags=re.IGNORECASE):
            val = tramite.get(col)
            if pd.notna(val) and str(val).strip():
                return _valor_celda_str(val)
    for col in tramite.index:
        val = tramite.get(col)
        if pd.notna(val) and re.match(r"^(19|20)\d{2}$", str(val).strip()):
            return str(val).strip()
    return ""


def _valor_codi_tramite(tramite) -> str:
    if "Unnamed: 5" in tramite.index:
        val = str(tramite.get("Unnamed: 5", "")).strip()
        if val in codigos_validacion:
            return val
    for col in tramite.index:
        val = str(tramite.get(col, "")).strip()
        if val in codigos_validacion:
            return val
    return ""


def sanitizar_nombre_archivo(nombre: str) -> str:
    nombre = unicodedata.normalize("NFD", str(nombre))
    nombre = "".join(c for c in nombre if unicodedata.category(c) != "Mn")
    nombre = re.sub(r"[^\w\s-]", "", nombre)
    nombre = nombre.strip().replace(" ", "_")
    return nombre


def _normalitzar_nom(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


def _get_tramite_valor(tramite, nom_logic, default=""):
    if nom_logic in tramite.index:
        return tramite[nom_logic]
    norm = _normalitzar_nom(nom_logic)
    for col in tramite.index:
        if _normalitzar_nom(str(col)) == norm:
            return tramite[col]
    return default


def crear_timestamp(fecha, hora):
    try:
        if isinstance(fecha, pd.Series) and isinstance(hora, pd.Series):
            hora_str = hora.apply(lambda x: x.strftime("%H:%M:%S") if hasattr(x, 'strftime') else str(x))
            return pd.to_datetime(fecha.astype(str) + " " + hora_str, errors="coerce")
        if hasattr(fecha, 'year') and hasattr(hora, 'hour'):
            return pd.Timestamp(fecha.year, fecha.month, fecha.day,
                                hora.hour, hora.minute, getattr(hora, 'second', 0))
        fecha_str = str(fecha).split()[0] if ' ' in str(fecha) else str(fecha)
        if hasattr(hora, 'strftime'):
            hora_str = hora.strftime("%H:%M:%S")
        else:
            hora_str = str(hora).split()[-1] if ' ' in str(hora) else str(hora)
        return pd.to_datetime(f"{fecha_str} {hora_str}", errors="coerce")
    except Exception:
        return pd.NaT


def tiempo_a_minutos(tiempo_str):
    if tiempo_str is None or (isinstance(tiempo_str, float) and pd.isna(tiempo_str)):
        return 0
    try:
        if hasattr(tiempo_str, "hour"):
            return tiempo_str.hour * 60 + tiempo_str.minute + tiempo_str.second / 60
        if isinstance(tiempo_str, pd.Timedelta):
            return tiempo_str.total_seconds() / 60
        if isinstance(tiempo_str, (int, float)):
            if 0 <= tiempo_str < 1:
                return tiempo_str * 24 * 60
            return 0
        s = str(tiempo_str).strip()
        partes = s.split(":")
        if len(partes) == 2:
            return int(partes[0]) * 60 + int(partes[1])
        if len(partes) == 3:
            return (int(partes[0]) * 60) + int(partes[1]) + (int(partes[2]) / 60)
        return 0
    except (ValueError, TypeError, AttributeError):
        return 0


def obtener_periodo_informe(df_visites: pd.DataFrame) -> tuple:
    def _periodo_per_defecte():
        hoy = datetime.datetime.now()
        mes_ingles = hoy.strftime("%B")
        mes_catalan_archivo = MESES_CATALAN.get(mes_ingles, mes_ingles)
        periodo_archivo = f"{hoy.strftime('%d')}_{hoy.strftime('%d')}_{mes_catalan_archivo}_{hoy.year}"
        texto_periodo = (f"Període del {hoy.strftime('%d')} al {hoy.strftime('%d')} "
                         f"{MESES_CATALAN_LONG.get(mes_ingles, mes_ingles)} del {hoy.year}")
        return hoy, hoy, periodo_archivo, texto_periodo

    if df_visites.empty or "Data Inici Visita" not in df_visites.columns:
        return _periodo_per_defecte()

    df_work = df_visites.copy()
    if not pd.api.types.is_datetime64_any_dtype(df_work["Data Inici Visita"]):
        df_work["Data Inici Visita"] = pd.to_datetime(df_work["Data Inici Visita"], errors="coerce")
    fecha_inicio = df_work["Data Inici Visita"].min()
    fecha_fin = df_work["Data Inici Visita"].max()
    if pd.isna(fecha_inicio) or pd.isna(fecha_fin):
        return _periodo_per_defecte()
    mes_ingles = fecha_fin.strftime("%B")
    mes_catalan_archivo = MESES_CATALAN.get(mes_ingles, mes_ingles)
    periodo_archivo = (f"{fecha_inicio.strftime('%d')}_{fecha_fin.strftime('%d')}"
                       f"_{mes_catalan_archivo}_{fecha_fin.year}")
    mes_catalan_mostrar = MESES_CATALAN_LONG.get(mes_ingles, mes_ingles)
    texto_periodo = (f"Període del {fecha_inicio.strftime('%d')} al {fecha_fin.strftime('%d')} "
                     f"{mes_catalan_mostrar} del {fecha_fin.year}")
    return fecha_inicio, fecha_fin, periodo_archivo, texto_periodo


# ============================================================================
# PROCESSAMENT (adaptat a bytes)
# ============================================================================

def procesar_visitas_bytes(data: bytes, log=print):
    """Carrega i filtra visites amb temps d'espera > mínim. Retorna (filtrat, original)."""
    df_temp = pd.read_excel(BytesIO(data), skiprows=FILAS_SKIP)

    _mapeo_v = MAPEO_COLUMNES_VISITES
    _rename_v = {v: k for k, v in _mapeo_v.items() if v and v != k}
    if _rename_v:
        df_temp = df_temp.rename(columns=_rename_v)

    for col in ["Data Inici Visita", "Data Atenció Visita", "Data Fi Visita"]:
        if col in df_temp.columns:
            df_temp[col] = pd.to_datetime(df_temp[col], errors="coerce")

    temps_col = None
    for col in df_temp.columns:
        if "Temps d'espera" in str(col):
            temps_col = col
            break
    if temps_col is None:
        raise KeyError("No s'ha trobat la columna 'Temps d'espera' a l'arxiu de visites")

    df_temp["minutos"] = df_temp[temps_col].apply(tiempo_a_minutos)
    df_visites = df_temp[df_temp["minutos"] > TIEMPO_ESPERA_MINIMO].copy().drop("minutos", axis=1)
    df_temp = df_temp.drop("minutos", axis=1)
    log(f"Visites amb espera > {TIEMPO_ESPERA_MINIMO} min: {len(df_visites)} (de {len(df_temp)})\n")
    return df_visites, df_temp


def procesar_entitats_bytes(data: bytes, log=print):
    """Carrega entitats i filtra segons codis de validació."""
    df_temp = pd.read_excel(BytesIO(data), skiprows=FILAS_SKIP)

    _mapeo_e = MAPEO_COLUMNES_ENTITATS
    _rename_e = {v: k for k, v in _mapeo_e.items() if v and v != k}
    if _rename_e:
        df_temp = df_temp.rename(columns=_rename_e)

    columna_codigo = None
    if "Unnamed: 5" in df_temp.columns:
        muestra = df_temp["Unnamed: 5"].astype(str).tolist()
        if sum(1 for x in muestra if x in codigos_validacion) > 0:
            columna_codigo = "Unnamed: 5"
    if not columna_codigo:
        candidatas = [c for c in df_temp.columns if df_temp[c].astype(str).str.len().mean() < 10]
        mejor, maxc = None, 0
        for col in candidatas:
            valores = df_temp[col].astype(str).tolist()
            coinc = sum(1 for v in valores if v in codigos_validacion)
            if coinc > maxc:
                maxc, mejor = coinc, col
        if mejor and maxc > 0:
            columna_codigo = mejor
    if not columna_codigo:
        for col in df_temp.columns:
            if "codi" in str(col).lower():
                columna_codigo = col
                break
    if not columna_codigo:
        raise KeyError("No s'ha pogut identificar la columna de codis de tràmit")

    df_entitats = df_temp[df_temp[columna_codigo].isin(codigos_validacion)].copy()
    if "Persona Física" in df_entitats.columns:
        df_entitats.loc[:, "Persona Física"] = df_entitats["Persona Física"].fillna(0)
    if "Unitat tramit (ui)" in df_entitats.columns:
        df_entitats.loc[:, "Unitat tramit (ui)"] = df_entitats["Unitat tramit (ui)"].astype(str).str.strip()
    for col in ["Data inici tramit"]:
        if col in df_entitats.columns:
            df_entitats.loc[:, col] = pd.to_datetime(df_entitats[col], errors="coerce")
    log(f"Tràmits vàlids (codis reconeguts): {len(df_entitats)}\n")
    return df_entitats


def procesar_profesionales(df_visites, df_entitats):
    try:
        df_visites = df_visites.copy()
        mask_prof = df_visites["Motiu Visita"] == "Professional"
        if "Persona Física" in df_entitats.columns:
            fichas_entitats = set(df_entitats["Persona Física"].dropna().unique())
        else:
            fichas_entitats = set()
        mask_in_entitats = df_visites["Fitxa SIP Entitat"].isin(fichas_entitats)
        fichas_reclasificadas = set(df_visites.loc[mask_prof & mask_in_entitats, "Fitxa SIP Entitat"].unique())
        fichas_confirmadas = set(df_visites.loc[mask_prof & ~mask_in_entitats, "Fitxa SIP Entitat"].unique())
        df_visites = df_visites.copy()
        df_visites.loc[mask_prof & mask_in_entitats, "Motiu Visita"] = "Particular"
        professional_df = df_visites.loc[mask_prof & ~mask_in_entitats].copy()
        cambios = ([{"ficha": f, "tipo": "reclasificación", "detalle": "De Professional a Particular"} for f in fichas_reclasificadas]
                   + [{"ficha": f, "tipo": "confirmación", "detalle": "Confirmado como professional"} for f in fichas_confirmadas])
        return df_visites, professional_df, cambios
    except (KeyError, ValueError) as e:
        print(f"Error al procesar professionals: {e}")
        return df_visites, pd.DataFrame(columns=df_visites.columns), []


def procesar_fichas_0(df_visites, professional_df, cambios_realizados):
    try:
        visitas_0 = df_visites[(df_visites["Fitxa SIP Entitat"] == 0) &
                               (df_visites["Motiu Visita"] != "Professional")].copy()
        if visitas_0.empty:
            return professional_df, cambios_realizados
        if professional_df is None or professional_df.empty:
            professional_df = visitas_0.copy()
        else:
            professional_df = pd.concat([professional_df, visitas_0], ignore_index=True)
        professional_df.loc[professional_df["Fitxa SIP Entitat"] == 0, "Motiu Visita"] = "Professional"
        for _, visita in visitas_0.iterrows():
            cambios_realizados.append({"ficha": visita["Fitxa SIP Entitat"], "tipo": "reclasificación",
                                       "detalle": f"De {visita['Motiu Visita']} a Professional (ficha 0)"})
        return professional_df, cambios_realizados
    except Exception as e:
        print(f"Error al procesar fitxes 0: {e}")
        return professional_df, cambios_realizados


def analizar_visita_tramites(visita, tramites_persona):
    try:
        entrada = crear_timestamp(visita["Data Inici Visita"], visita["Hora Inici Visita"])
        atencion = crear_timestamp(visita["Data Atenció Visita"], visita["Hora atenció visita"])
        salida = crear_timestamp(visita["Data Fi Visita"], visita["Hora Fi Visita"])
        if pd.isna(entrada) or pd.isna(atencion) or pd.isna(salida):
            return None
        limite_tiempo = atencion - pd.Timedelta(minutes=MINUTOS_LIMITE_TRAMITE)
        if tramites_persona.empty:
            return None
        tramites_persona = tramites_persona.copy()
        tramites_persona["timestamp"] = crear_timestamp(
            tramites_persona["Data inici tramit"], tramites_persona["Hora inici tràmit"])
        tramites_validos = tramites_persona[
            (tramites_persona["timestamp"] >= entrada) &
            (tramites_persona["timestamp"] <= salida) &
            (tramites_persona["timestamp"] <= limite_tiempo)]
        if tramites_validos.empty:
            return None
        return tramites_validos.sort_values("timestamp").iloc[0]
    except Exception as e:
        print(f"Error al analizar visita-tramits: {e}")
        return None


def procesar_validaciones(df_visites, df_entitats, progress=None):
    try:
        if df_visites.empty or df_entitats.empty:
            return []
        if "Persona Física" not in df_entitats.columns:
            return []
        entitats_por_persona = df_entitats.groupby("Persona Física")
        validats = []
        total = len(df_visites)
        for i, (_, visita) in enumerate(df_visites.iterrows(), 1):
            if progress is not None and (i % 25 == 0 or i == total):
                progress.set_step("Validant visites...")
                progress.set_progress(i, total)
            ficha = visita["Fitxa SIP Entitat"]
            if ficha not in entitats_por_persona.groups:
                continue
            tramites_persona = entitats_por_persona.get_group(ficha)
            tramite_validador = analizar_visita_tramites(visita, tramites_persona)
            if tramite_validador is not None:
                validats.append({"visita": visita, "tramite": tramite_validador})
        if progress is not None:
            progress.set_progress(total, total)
        return validats
    except Exception as e:
        print(f"Error al procesar validacions: {e}")
        return []


def procesar_validaciones_final(validats, professional_df):
    columnas = [
        "Tipus", "Fitxa SIP Entitat", "Motiu Visita", "Instal·lació",
        "Data Inici Visita", "Hora Inici Visita", "Data Atenció Visita",
        "Hora atenció visita", "Data Fi Visita", "Hora Fi Visita",
        "Temps d'espera", "Procediment", "Any", "Tipus tràmit",
        "Codi tràmit", "Data inici tramit", "Hora inici tràmit",
        "Agent instructor tràmit", "Unitat tràmit"
    ]
    registros = []
    for validacion in validats:
        visita = validacion["visita"]
        tramite = validacion["tramite"]
        procediment = _get_tramite_valor(tramite, "Procediment", "")
        any_tramit = _valor_any_tramite(tramite)
        tipus_tramit = _get_tramite_valor(tramite, "Tipus tramit", "")
        codi_tramit = _valor_codi_tramite(tramite)
        data_inici_tramit = _get_tramite_valor(tramite, "Data inici tramit", None)
        hora_inici_tramit = _get_tramite_valor(tramite, "Hora inici tràmit", None)
        agent_tramit = _get_tramite_valor(tramite, "Agent instructor tramit", "")
        _ut = _get_tramite_valor(tramite, "Unitat tramit (ui)", None)
        unitat_tramit = str(_ut).strip() if _ut is not None and pd.notna(_ut) else ""
        registros.append({
            "Tipus": "Registre Positiu", "Fitxa SIP Entitat": visita["Fitxa SIP Entitat"],
            "Motiu Visita": visita["Motiu Visita"], "Instal·lació": visita["Instal·lació"],
            "Data Inici Visita": visita["Data Inici Visita"], "Hora Inici Visita": visita["Hora Inici Visita"],
            "Data Atenció Visita": visita["Data Atenció Visita"], "Hora atenció visita": visita["Hora atenció visita"],
            "Data Fi Visita": visita["Data Fi Visita"], "Hora Fi Visita": visita["Hora Fi Visita"],
            "Temps d'espera": visita.get("Temps d'espera"), "Procediment": procediment, "Any": any_tramit,
            "Tipus tràmit": tipus_tramit, "Codi tràmit": codi_tramit, "Data inici tramit": data_inici_tramit,
            "Hora inici tràmit": hora_inici_tramit, "Agent instructor tràmit": agent_tramit, "Unitat tràmit": unitat_tramit,
        })
    if professional_df is not None and not professional_df.empty:
        for _, visita in professional_df.iterrows():
            registros.append({
                "Tipus": "Confirmació Professional", "Fitxa SIP Entitat": visita["Fitxa SIP Entitat"],
                "Motiu Visita": visita["Motiu Visita"], "Instal·lació": visita.get("Instal·lació", ""),
                "Data Inici Visita": visita.get("Data Inici Visita"), "Hora Inici Visita": visita.get("Hora Inici Visita"),
                "Data Atenció Visita": visita.get("Data Atenció Visita"), "Hora atenció visita": visita.get("Hora atenció visita"),
                "Data Fi Visita": visita.get("Data Fi Visita"), "Hora Fi Visita": visita.get("Hora Fi Visita"),
                "Temps d'espera": visita.get("Temps d'espera"), "Procediment": "", "Any": "",
                "Tipus tràmit": "", "Codi tràmit": "", "Data inici tramit": None, "Hora inici tràmit": None,
                "Agent instructor tràmit": "", "Unitat tràmit": "",
            })
    return pd.DataFrame(registros, columns=columnas)


# ============================================================================
# GENERACIÓ D'INFORMES (retornen bytes)
# ============================================================================

def insertar_logo(ws):
    if not LOGO_PATH:
        return False
    try:
        import os
        if not os.path.exists(LOGO_PATH) or os.path.getsize(LOGO_PATH) == 0:
            return False
        img = OpenpyxlImage(LOGO_PATH)
        max_height = 60
        ratio = (max_height / img.height) if (img.height and img.height > 0) else 1
        img.height = max_height
        img.width = int(img.width * ratio) if img.width and img.width > 0 else 150
        img.anchor = "B1"
        ws.add_image(img)
        ws.row_dimensions[1].height = max_height * 0.75
        return True
    except Exception as e:
        print(f"Advertència logo: {e}")
        return False


def _crear_estils_informe():
    s = ESTIL_INFORMES
    return {
        "fill_par": PatternFill(start_color=s["color_fila_par"], end_color=s["color_fila_par"], fill_type="solid"),
        "fill_hora": PatternFill(start_color=s["color_hora_modificar"], end_color=s["color_hora_modificar"], fill_type="solid"),
        "fill_prof": PatternFill(start_color=s["color_professional"], end_color=s["color_professional"], fill_type="solid"),
        "fill_header": PatternFill(start_color=s["color_encabezado"], end_color=s["color_encabezado"], fill_type="solid"),
        "font_datos": Font(name="Arial", size=s["fuente_datos"]),
        "font_hora": Font(name="Arial", size=s["fuente_datos"], bold=True, color="C65911"),
        "font_prof": Font(name="Arial", size=s["fuente_datos"], bold=True, color="548235"),
        "font_header": Font(name="Arial", size=s["fuente_encabezado"], bold=True, color=s["color_texto_encabezado"]),
        "font_titulo": Font(name="Arial", size=s["fuente_titulo"], bold=True, color=s["color_titulo"]),
        "font_subtitulo": Font(name="Arial", size=s["fuente_subtitulo"], bold=True, color="505050"),
        "align_cc": Alignment(horizontal="center", vertical="center"),
        "align_cc_wrap": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border_thin": Border(left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
                              top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0")),
        "border_header": Border(left=Side(style="medium", color="1F4E78"), right=Side(style="medium", color="1F4E78"),
                                top=Side(style="medium", color="1F4E78"), bottom=Side(style="medium", color="1F4E78")),
    }


def _escriure_fila_principal(ws, fila, num_ordre, data, comisaria, es_fila_par, st, col_offset=0):
    s_inf = ESTIL_INFORMES
    es_professional = data["Motiu Visita"] == "Professional"
    ws.row_dimensions[fila].height = s_inf["altura_datos"]
    col = 2

    cell = ws.cell(row=fila, column=col, value=num_ordre)
    cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_fila_par: cell.fill = st["fill_par"]
    col += 1

    cell = ws.cell(row=fila, column=col, value=data["Data Inici Visita"])
    cell.number_format = "DD/MM/YYYY"
    cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_fila_par: cell.fill = st["fill_par"]
    col += 1

    if col_offset == 1:
        cell = ws.cell(row=fila, column=col, value=data["Instal·lació"])
        cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
        if es_fila_par: cell.fill = st["fill_par"]
        col += 1

    cell = ws.cell(row=fila, column=col, value=MAPEO_COMISARIAS.get(comisaria, comisaria))
    cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_fila_par: cell.fill = st["fill_par"]
    col += 1

    cell = ws.cell(row=fila, column=col, value=data["Fitxa SIP Entitat"])
    cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_fila_par: cell.fill = st["fill_par"]
    col += 1

    cell = ws.cell(row=fila, column=col, value=data["Hora atenció visita"])
    cell.number_format = "[h]:mm"
    cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_fila_par: cell.fill = st["fill_par"]
    col += 1

    if es_professional:
        hora_modificar = data["Hora Inici Visita"]
        observacions = "Professional"
    else:
        hora_modificar = data.get("Hora inici tràmit") if pd.notna(data.get("Hora inici tràmit")) else ""
        observacions = ""

    cell = ws.cell(row=fila, column=col, value=hora_modificar)
    cell.number_format = "[h]:mm"
    cell.font = st["font_hora"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    cell.fill = st["fill_hora"]
    col += 1

    procediment = _valor_celda_str(data.get("Procediment"))
    any_tramit = _valor_celda_str(data.get("Any"))
    diligencies = f"{procediment}/{any_tramit}" if procediment and any_tramit else ""
    cell = ws.cell(row=fila, column=col, value=diligencies)
    cell.font = st["font_datos"]; cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_fila_par: cell.fill = st["fill_par"]
    col += 1

    cell = ws.cell(row=fila, column=col, value=observacions)
    cell.alignment = st["align_cc"]; cell.border = st["border_thin"]
    if es_professional:
        cell.fill = st["fill_prof"]; cell.font = st["font_prof"]
    elif es_fila_par:
        cell.fill = st["fill_par"]; cell.font = st["font_datos"]
    else:
        cell.font = st["font_datos"]


def _escriure_encapçalaments(ws, fila, headers, st):
    for col, header in enumerate(headers, 2):
        cell = ws.cell(row=fila, column=col, value=header)
        cell.font = st["font_header"]; cell.fill = st["fill_header"]
        cell.alignment = st["align_cc_wrap"]; cell.border = st["border_header"]
    ws.row_dimensions[fila].height = ESTIL_INFORMES["altura_encabezado"]


def _wb_to_bytes(wb):
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def guardar_excel_validacion_bytes(df_visites_original, proceso_validacion_df,
                                   visites_bytes, entitats_bytes, progress=None):
    """Construeix el consolidat (Visites ORIGINAL + Entitats ORIGINAL + Registre Validat). Retorna bytes."""
    wb_visitas = load_workbook(BytesIO(visites_bytes))
    wb_entitats = load_workbook(BytesIO(entitats_bytes))
    wb_nuevo = Workbook()
    wb_nuevo.remove(wb_nuevo.active)
    try:
        hoja_visitas = wb_visitas.active
        nueva_v = wb_nuevo.create_sheet(title="Visites ORIGINAL")
        for row in hoja_visitas.iter_rows():
            nueva_v.append([cell.value for cell in row])

        hoja_entitats = wb_entitats.active
        nueva_e = wb_nuevo.create_sheet(title="Entitats ORIGINAL")
        for row in hoja_entitats.iter_rows():
            nueva_e.append([cell.value for cell in row])

        nueva_val = wb_nuevo.create_sheet(title="Registre Validat")
        df_ord = proceso_validacion_df.sort_values(by=["Instal·lació", "Data Inici Visita", "Hora Inici Visita"])
        nueva_val.append(list(df_ord.columns))
        for _, row in df_ord.iterrows():
            nueva_val.append(row.tolist())

        wb_visitas.close(); wb_entitats.close()
        wb_visitas = wb_entitats = None

        for sheet in wb_nuevo.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        sheet = nueva_val
        anchos = {"A": 23, "B": 15, "C": 12, "D": 29, "E": 17, "F": 16, "G": 18, "H": 16, "I": 12,
                  "J": 12, "K": 15, "L": 12, "M": 8, "N": 47, "O": 10, "P": 15, "Q": 17, "R": 22, "S": 53}
        for col_letter, ancho in anchos.items():
            sheet.column_dimensions[col_letter].width = ancho
        for col in ["E", "G", "I", "P"]:
            for cell in sheet[col][1:]:
                if cell.value:
                    cell.number_format = "dd/mm/yyyy"; cell.alignment = Alignment(horizontal="center")
        for col in ["F", "H", "J", "Q", "K"]:
            for cell in sheet[col][1:]:
                if cell.value:
                    cell.number_format = "[h]:mm"; cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        return _wb_to_bytes(wb_nuevo)
    finally:
        for wb in (wb_visitas, wb_entitats):
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass


def generar_informes_por_comisaria_bytes(proceso_validacion_df, df_visites_original, progress=None):
    """Retorna llista de (nom_fitxer, bytes), un informe per comissaria."""
    s = ESTIL_INFORMES
    st = _crear_estils_informe()
    _, _, periodo_archivo, texto_periodo = obtener_periodo_informe(df_visites_original)
    comisarias = proceso_validacion_df["Instal·lació"].unique()
    enc = ["Ordre", "Data", "ABP Sol·licitant", "Fitxa SIP", "Hora atenció visita",
           "Hora a Modificar", "Diligències", "OBSERVACIONS"]
    resultats = []
    total = len(comisarias)
    for comi_idx, comisaria in enumerate(comisarias, 1):
        if pd.isna(comisaria):
            continue
        comisaria = str(comisaria)
        if progress is not None and (comi_idx == 1 or comi_idx % 5 == 0 or comi_idx == total):
            progress.set_step(f"Informe: {comisaria[:40]}...")
            progress.set_progress(comi_idx, total)
        df_com = proceso_validacion_df[proceso_validacion_df["Instal·lació"] == comisaria].copy().sort_values(
            by=["Data Inici Visita", "Hora Inici Visita"], ascending=True)
        if df_com.empty:
            continue
        wb = Workbook(); ws = wb.active; ws.title = comisaria[:31]
        insertar_logo(ws)
        ws.merge_cells("B2:I2")
        ws["B2"] = f"Modificacions del Llibre de Visites de {comisaria}"
        ws["B2"].font = st["font_titulo"]; ws["B2"].alignment = st["align_cc"]
        ws.row_dimensions[2].height = s["altura_titulo"]
        ws.merge_cells("B4:I4")
        ws["B4"] = texto_periodo
        ws["B4"].font = st["font_subtitulo"]; ws["B4"].alignment = st["align_cc"]
        ws.row_dimensions[5].height = 5
        _escriure_encapçalaments(ws, 7, enc, st)
        ws.freeze_panes = "B8"
        ws.auto_filter.ref = f"B7:I{len(df_com) + 7}"
        for i, (_, data) in enumerate(df_com.iterrows(), 1):
            _escriure_fila_principal(ws, i + 7, i, data, comisaria, i % 2 == 0, st, col_offset=0)
        for c, w in {"B": 12, "C": 14, "D": 26, "E": 16, "F": 26, "G": 26, "H": 26, "I": 26}.items():
            ws.column_dimensions[c].width = w
        nom = f"Informe_{sanitizar_nombre_archivo(comisaria)}_{periodo_archivo}.xlsx"
        resultats.append((nom, _wb_to_bytes(wb)))
        wb.close()
    return resultats


def generar_informe_consolidado_bytes(proceso_validacion_df, df_visites_original, progress=None):
    """Retorna (nom_fitxer, bytes) amb totes les comissaries."""
    s = ESTIL_INFORMES
    st = _crear_estils_informe()
    _, _, periodo_archivo, texto_periodo = obtener_periodo_informe(df_visites_original)
    df_cons = proceso_validacion_df.sort_values(by=["Instal·lació", "Data Inici Visita", "Hora Inici Visita"])
    enc = ["Ordre", "Data", "Comissaria", "ABP Sol·licitant", "Fitxa SIP",
           "Hora atenció visita", "Hora a Modificar", "Diligències", "OBSERVACIONS"]
    wb = Workbook(); ws = wb.active; ws.title = "Totes les comissaries"
    insertar_logo(ws)
    ws.merge_cells("B2:J2")
    ws["B2"] = "Modificacions del Llibre de Visites (Totes les comissaries)"
    ws["B2"].font = st["font_titulo"]; ws["B2"].alignment = st["align_cc"]
    ws.row_dimensions[2].height = s["altura_titulo"]
    ws.merge_cells("B4:J4")
    ws["B4"] = texto_periodo
    ws["B4"].font = st["font_subtitulo"]; ws["B4"].alignment = st["align_cc"]
    ws.row_dimensions[5].height = 5
    _escriure_encapçalaments(ws, 7, enc, st)
    ws.freeze_panes = "B8"
    ws.auto_filter.ref = f"B7:J{len(df_cons) + 7}"
    total = len(df_cons)
    for i, (_, data) in enumerate(df_cons.iterrows(), 1):
        if progress is not None and (i % 50 == 0 or i == total):
            progress.set_step("Informe consolidat...")
            progress.set_progress(i, total)
        _escriure_fila_principal(ws, i + 7, i, data, data["Instal·lació"], i % 2 == 0, st, col_offset=1)
    for c, w in {"B": 12, "C": 14, "D": 32, "E": 26, "F": 16, "G": 26, "H": 26, "I": 26, "J": 26}.items():
        ws.column_dimensions[c].width = w
    nom = f"Informe_Totes_Comissaries_{periodo_archivo}.xlsx"
    return nom, _wb_to_bytes(wb)


# ============================================================================
# ORQUESTRADOR (entrada: 2 bytes; sortida: ZIP en bytes)
# ============================================================================

class _Progress:
    """Adaptador perquè les funcions internes informin del progrés a la web."""
    def __init__(self, log_cb=None, progress_cb=None):
        self._log = log_cb
        self._prog = progress_cb
    def set_step(self, text):
        if self._log:
            self._log(str(text) + "\n")
    def set_progress(self, current, total=None):
        if self._prog and total:
            self._prog(current, total)
    def close(self):
        pass


def processa_auditoria(visites_bytes, entitats_bytes, opcion_informe="ambdos",
                       log_cb=None, progress_cb=None):
    """Executa tot el procés d'auditoria. Retorna dict amb el ZIP (bytes) i estadístiques.

    opcion_informe: 'ambdos' | 'individuals' | 'consolidat' | 'cap'
    """
    def _log(m):
        if log_cb:
            log_cb(m)

    progress = _Progress(log_cb, progress_cb)
    visites_bytes = bytes(visites_bytes)
    entitats_bytes = bytes(entitats_bytes)

    _log("Processant visites i entitats...\n")
    df_visites_filtrat, df_visites_original = procesar_visitas_bytes(visites_bytes, log=_log)
    df_entitats = procesar_entitats_bytes(entitats_bytes, log=_log)

    if df_visites_filtrat.empty:
        raise ValueError("No s'han trobat visites amb temps d'espera superior al mínim configurat.")
    if df_entitats.empty:
        raise ValueError("No s'han trobat tràmits vàlids a l'arxiu d'entitats.")

    _log("Processant professionals...\n")
    df_visites_prof, professional_df, canvis = procesar_profesionales(df_visites_filtrat, df_entitats)
    professional_df, canvis = procesar_fichas_0(df_visites_prof, professional_df, canvis)

    _log("Validant visites...\n")
    validats = procesar_validaciones(df_visites_prof, df_entitats, progress=progress)
    proceso_validacion_df = procesar_validaciones_final(validats, professional_df)

    if proceso_validacion_df.empty:
        raise ValueError("No s'han trobat registres per validar amb els criteris actuals.")

    _, _, periodo_archivo, _ = obtener_periodo_informe(df_visites_original)

    fitxers = {}
    _log("Generant consolidat (Registre validat)...\n")
    fitxers[f"Registre_validat_{periodo_archivo}.xlsx"] = guardar_excel_validacion_bytes(
        df_visites_original, proceso_validacion_df, visites_bytes, entitats_bytes, progress=progress)

    if opcion_informe in ("ambdos", "individuals"):
        _log("Generant informes per comissaria...\n")
        for nom, b in generar_informes_por_comisaria_bytes(proceso_validacion_df, df_visites_original, progress=progress):
            fitxers[nom] = b
    if opcion_informe in ("ambdos", "consolidat"):
        _log("Generant informe consolidat...\n")
        nom, b = generar_informe_consolidado_bytes(proceso_validacion_df, df_visites_original, progress=progress)
        fitxers[nom] = b

    # Empaquetar en ZIP
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nom, b in fitxers.items():
            zf.writestr(nom, b)
    _log(f"Generats {len(fitxers)} fitxers, empaquetats en ZIP.\n")

    n_positius = int((proceso_validacion_df["Tipus"] == "Registre Positiu").sum())
    n_prof = int((proceso_validacion_df["Tipus"] == "Confirmació Professional").sum())
    n_comissaries = int(proceso_validacion_df["Instal·lació"].nunique())
    stats = {
        "visites_filtrades": len(df_visites_filtrat),
        "tramits_valids": len(df_entitats),
        "registres_validats": len(proceso_validacion_df),
        "registres_positius": n_positius,
        "confirmacions_professional": n_prof,
        "comissaries": n_comissaries,
        "num_fitxers": len(fitxers),
        "data_proces": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }
    return {"zip": zip_buf.getvalue(), "stats": stats, "noms_fitxers": list(fitxers.keys())}


def resum_text(stats):
    return "\n".join([
        "=" * 40, "RESUM DE L'AUDITORIA", "=" * 40,
        f"Visites amb espera > mínim: {stats['visites_filtrades']}",
        f"Tràmits vàlids: {stats['tramits_valids']}",
        f"Registres validats: {stats['registres_validats']}",
        f"  - Registres positius: {stats['registres_positius']}",
        f"  - Confirmacions professional: {stats['confirmacions_professional']}",
        f"Comissaries: {stats['comissaries']}",
        f"Fitxers generats: {stats['num_fitxers']}",
        "=" * 40,
    ])
