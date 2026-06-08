#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CORREGIDOR D'ÀMBITS - NUCLI WEB (Pyodide)
USC Manresa - Mossos d'Esquadra

Lògica de correcció d'àmbits EXTRETA de corregidor.py (v6.4), SENSE Tkinter ni
PIL, perquè pugui executar-se dins del navegador amb Pyodide (WebAssembly).

La classe CorregidorAmbits i totes les regles són IDÈNTIQUES a la versió
d'escriptori ja validada. Només canvia la capa d'entrada/sortida:
  - L'Excel d'entrada arriba com a `bytes` (no com a ruta de fitxer).
  - L'Excel de sortida es retorna com a `bytes` (per descarregar-lo al navegador).
  - La configuració s'injecta amb set_config() (no es llegeix de config.json).

Cap dada surt del navegador: tot el processament passa a la màquina de l'usuari.
"""

import re
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple, Any, Set
from dataclasses import dataclass

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook


VERSION = "6.4"

# Configuració injectada des del JavaScript (secció 'corregidor' del config.json).
CONFIG: Dict[str, Any] = {}


def set_config(cfg: Dict[str, Any]) -> None:
    """Injecta la configuració (cridat des del navegador abans de processar)."""
    global CONFIG
    CONFIG = cfg or {}


COLUMNES_ALIAS = {
    "nom vial del fet": "Nom Vial del fet",
    "nom viale del fet": "Nom Vial del fet",
}
COL_UNITAT = "Unitat (ui)"


def _normalitza_nom_columna(nom: str) -> str:
    if not nom or not isinstance(nom, str):
        return ""
    return " ".join(str(nom).strip().split())


def _aplica_alias_columnes(df: pd.DataFrame) -> pd.DataFrame:
    renoms = {}
    for col in df.columns:
        norm = _normalitza_nom_columna(str(col)).lower()
        if norm in COLUMNES_ALIAS:
            renoms[col] = COLUMNES_ALIAS[norm]
    if renoms:
        df = df.rename(columns=renoms)
    return df


def _get_config(key_path: str, default: Any) -> Any:
    keys = key_path.split(".")
    val: Any = CONFIG
    for k in keys:
        if not isinstance(val, dict) or k not in val:
            return default
        val = val[k]
    return val


def _obte_unitats_filtrar() -> Optional[Set[str]]:
    """Retorna None si processar totes; set(noms) si filtrar per unitats específiques."""
    unitats = _get_config("fitxers.unitats_filtrar", [])
    if not unitats or not isinstance(unitats, list):
        return None
    sel = {str(u).strip() for u in unitats if u and str(u).strip()}
    return sel if sel else None


# ============================================================================
# MAPA DE COLORS PER ÀMBIT
# ============================================================================

AMBIT_COLOR_MAP = {
    "Sense àmbit": "A6A6A6",
    "Estafes": "FFC000",
    "POE Habitatge": "00B0F0",
    "POE Comerç i Empresa": "70AD47",
    "POE Oci nocturn": "FF8C00",
    "POE Pista": "2E5C8A",
    "POE Centres educatius": "D97634",
    "POE Món rural": "9DC3E6",
    "POE LCD Metall": "B4C7E7",
    "POE Tràfic d'Éssers Humans": "F4B084",
    "INTERNET": "FFFF99",
    "Violència de gènere": "FFFF99",
    "Violència domèstica": "FFFF99",
    "Violència sexual": "FFFF99",
    "Violència filio parental": "FFFF99",
    "Altres violències masclistes": "FFFF99",
    "Medi ambient": "E8F4F8",
    "ODI I DISCRIMINACIÓ": "F0E8F4",
    "DIVERSOS": "F4F0E8",
    "Entorn laboral": "E8F4E8",
}


def get_color_by_ambit(ambit_text: str) -> str:
    if not ambit_text:
        return "D3D3D3"
    return AMBIT_COLOR_MAP.get(str(ambit_text).strip(), "D3D3D3")


DEFAULT_COLUMNES_OBLIGATORIES = [
    "Ambit procediment",
    "Tipus fet (nivell 2)",
    "Tipus punt",
    "Motiu resolució",
    "Procediment",
    "Data Inici (fet)",
    "Unitat (ui)",
]

COLUMNES_RECOMANADES = [
    "Hora (franja) Inici (fet)",
    "Nom Vial del fet",
    "Tipus vial",
]


def _col_canonica(nom: str) -> str:
    norm = _normalitza_nom_columna(str(nom)).lower()
    return COLUMNES_ALIAS.get(norm, nom)


def _valida_columnes(df: pd.DataFrame) -> Tuple[bool, List[str]]:
    columnes_oblig = _get_config("excel.columnes_obligatories", []) or DEFAULT_COLUMNES_OBLIGATORIES
    df_cols = set(df.columns)
    faltants = []
    for c in columnes_oblig:
        canonica = _col_canonica(c)
        if c in df_cols or canonica in df_cols:
            continue
        faltants.append(c)
    return len(faltants) == 0, faltants


def _avisa_columnes_recomanades(df: pd.DataFrame, log_cb=None) -> None:
    faltants = [c for c in COLUMNES_RECOMANADES if c not in df.columns]
    if faltants and log_cb:
        log_cb(f"Columnes recomanades no trobades: {', '.join(faltants)}\n")


def _format_procediment(registre: Dict) -> str:
    proc = registre.get("Procediment", "")
    data_val = registre.get("Data Inici (fet)", "")
    proc_str = ""
    if pd.notna(proc) and str(proc).strip():
        try:
            proc_str = str(int(float(proc)))
        except (ValueError, TypeError):
            proc_str = str(proc).strip()
    year_str = ""
    if pd.notna(data_val):
        try:
            year_str = str(pd.to_datetime(data_val).year)
        except (ValueError, TypeError):
            pass
    return f"{proc_str}/{year_str}" if proc_str and year_str else proc_str


# ============================================================================
# DEFINICIONS DE DADES I REGLES  (idèntic a corregidor.py)
# ============================================================================


@dataclass
class ReglaCorreccio:
    id_regla: str
    prioritat: str
    descripcio: str
    ambit_desti: str
    motiu: str
    categoria: str
    color_original: str
    color_suggerit: str


class CorregidorAmbits:
    HORA_REGEX = re.compile(r"(\d{1,2})")
    ART_249_REGEX = re.compile(r"\b249\b|art\.\s*249", re.IGNORECASE)

    def __init__(self, log_cb=None, regles_desactivades=None):
        self._log_cb = log_cb
        self.regles_desactivades = set(regles_desactivades or ())
        self.ambits_protegits = {
            "Violència de gènere", "Violència domèstica", "Violència sexual",
            "Violència filio parental", "Altres violències masclistes",
            "Altres violencies masclistes", "Estafes",
        }
        self.ambits_valids = {
            "INTERNET", "POE Comerç i Empresa", "POE Habitatge",
            "POE Centres educatius", "POE Oci nocturn", "POE Món rural",
            "POE LCD Metall", "POE Pista", "POE Tràfic d'Éssers Humans",
            "Violència de gènere", "Violència domèstica", "Violència sexual",
            "Violència filio parental", "Estafes", "Sense àmbit",
            "Medi ambient", "ODI I DISCRIMINACIÓ", "DIVERSOS", "Entorn laboral",
            "Activitats delictives", "Atenci a la vctima", "Bullying menors dedat",
            "Discapacitat", "Persona vulnerable", "Transtorns mentalspsquics",
            "Altres violències masclistes", "Altres violencies masclistes",
            "Matrimoni forat", "Mutilacio genital femenina", "Nadons sostrets al nixer",
            "Afectacio a menors", "AmenacesCoaccionsExtorsions",
            "Assetjament major detat", "Atacs informatics", "Continguts nocius",
            "Incitacio a lodi i a la violencia", "Propietat intellectual",
            "Usurpacio de la identitat", "CongressosConvencions", "Creuers",
            "Esportiu", "Festes populars", "ManifestacionsConcentracions",
            "Penitenciari", "Signum", "LaboralTreball", "Mn del taxi i afins",
            "AntiguitatsObres dart", "Bandes llatines juvenils", "Grups urbans",
            "Altres", "Antigitanisme", "Antisemitisme", "Aporofbia",
            "Discapacitat fsicsensorial", "Discapacitat intellectualmental",
            "Edatisme", "tnicOrigen nacionalOrigen racial", "Islamofbia",
            "LGTBI-fbia", "Malaltia", "Orientacio politica", "Religis",
            "Sexisme", "Situacio familiar",
        }
        self.motius_exclusio = {
            "inexi. il.licit penal", "arxivat per duplicat",
            "incompetència territorial", "fets no constitutius de delicte",
        }
        self.vies_poe_pista = {
            "ap-7", "ap7", "ap-2", "ap2", "c-16", "c16", "c-32", "c32",
            "c-58", "c58", "a-2", "a2", "c-25", "c25",
        }
        self.tipus_punt_habitatge_valids = {
            "hab. 1a res. pis/apartament", "hab. 1a res. casa adossada",
            "hab. 1a res. casa aïllada/4vents", "hab. 1a res. masia/casa de camp",
            "hab. 2a res. pis/apartament", "hab. 2a res. casa adossada",
            "hab. 2a res. casa aïllada/4vents", "hab. 2a res. masia/casa de camp",
        }
        self.tipus_punt_habitatge_exclos = {
            "habitatge sense ús", "hab. pàrquing comunitari", "hab. pàrquing",
            "hab. traster", "hab. escala/portal finca/replà", "hab. altres espais",
            "hab. ascensor", "hab. garatge", "hab. jardí comunitari", "hab. jardí",
            "hab. pati comunitari", "hab. pati de finca", "hab. pati",
        }
        self.estafes_informatiques = {
            "estafes informàtiques art. 249", "estafa informàtica", "phishing",
            "smishing", "vishing", "credencials", "accés informàtic",
            "pesca de credencials", "art. 249.1a", "art. 249.2a",
            "art. 249 1 a", "art. 249 2 a",
        }
        self.estafes_targetes = {
            "estafes bancàries targetes art. 249", "targetes art. 249",
            "targeta sostreta", "targeta perduda", "ús no consentit targeta",
            "art. 249.1b", "art. 249.2b", "art. 249 1 b", "art. 249 2 b",
            "targetes bancàries",
        }
        self.estafes_ciberengany = {
            "fals familiar", "ceo fraud", "business email", "bec",
            "estafador amor", "estafador de l'amor", "falsa oferta treball",
            "suplantació identitat", "venedor online", "adquiridor online",
            "compra online", "venda online", "llogater online", "inversor online",
            "bizum", "càrrec targeta", "càrrec compte", "càrrec bancari",
            "transferència no autoritzada", "alta servei a nom",
        }
        self.estafes_basiques = {
            "consumir sense pagar", "combustible sense pagar", "restaurant sense pagar",
            "hotel sense pagar", "transport sense pagar", "canvi etiquetes",
            "scalping", "caixer", "distracció caixer", "falsos operaris",
            "gas llum telefonia", "rip deal", "tocomocho", "tocomotxo",
            "estampeta", "ponzi", "piramidal", "xec fals", "pagaré fals",
            "esquema piramidal",
        }
        self.regles = self._carrega_totes_regles()
        if self._log_cb:
            self._log_cb(
                f"Corregidor inicialitzat amb {len(self.regles)} regles | "
                f"Pauta Estafes GENER 2026 (v{VERSION})\n"
            )

    def _detecta_tipus_estafa(self, tipus_fet: str) -> Optional[str]:
        if not tipus_fet:
            return None
        text = str(tipus_fet).lower()
        if any(excl in text for excl in ["lleu", "menor", "inferior"]):
            if "estafa" not in text and "art. 248" not in text and "art. 249" not in text:
                return None
        if any(term in text for term in self.estafes_informatiques):
            return "estafa_informatica"
        if any(term in text for term in self.estafes_targetes):
            return "estafa_targetes"
        if any(term in text for term in self.estafes_ciberengany):
            return "estafa_ciberengany"
        if any(term in text for term in self.estafes_basiques):
            return "estafa_basica"
        if "estafa" in text or "art. 248" in text or "art. 251" in text:
            return "estafa_generica"
        if self.ART_249_REGEX.search(text):
            if "1 b" in text or "2 b" in text or "targeta" in text:
                return "estafa_targetes"
            if "1 a" in text or "2 a" in text or "informàtic" in text:
                return "estafa_informatica"
            return "estafa_generica"
        return None

    def _carrega_totes_regles(self) -> List[ReglaCorreccio]:
        return [
            ReglaCorreccio("estafa_informatica_estafes", "ALTA", "Estafes informàtiques → Estafes", "Estafes",
                "Estafa informàtica art. 249.1a) - Phishing/Smishing/Vishing/Credencials",
                "Estafes Ciberfacilitadores - Pesca", "FFFF99", "FFC000"),
            ReglaCorreccio("estafa_targetes_estafes", "ALTA", "Estafes targetes bancàries → Estafes", "Estafes",
                "Estafa targetes art. 249.1b) - Ús no consentit targeta bancària",
                "Estafes Ciberfacilitadores - Targetes", "FFFF99", "FFC000"),
            ReglaCorreccio("estafa_ciberengany_estafes", "ALTA", "Estafes ciberengany → Estafes", "Estafes",
                "Estafa ciberfacilitadora art. 248 - Suplantació/Transaccions/Càrrecs online",
                "Estafes Ciberfacilitadores - Engany Digital", "FFFF99", "FFC000"),
            ReglaCorreccio("estafa_basica_estafes", "MITJANA", "Estafes bàsiques → Estafes", "Estafes",
                "Estafa bàsica art. 248 - Mètodes tradicionals no digitals", "Estafes Bàsiques", "FFFF99", "FFC000"),
            ReglaCorreccio("estafa_generica_estafes", "MITJANA", "Altres estafes → Estafes", "Estafes",
                "Altres estafes art. 248-251 bis", "Estafes Genèriques", "FFFF99", "FFC000"),
            ReglaCorreccio("oci_nocturn_poe", "MITJANA", "Oci nocturn → POE Oci nocturn", "POE Oci nocturn",
                "Delicte en horari nocturn a establiment d'oci", "POE Oci Nocturn", "FFE699", "FF8C00"),
            ReglaCorreccio("patrimoni_comerc_poe", "MITJANA", "Delicte en comerç → POE Comerç", "POE Comerç i Empresa",
                "Delicte patrimonial en establiment comercial", "POE Patrimoni", "C6E0B4", "70AD47"),
            ReglaCorreccio("patrimoni_habitatge_poe", "MITJANA", "Delicte en habitatge → POE Habitatge", "POE Habitatge",
                "Delicte patrimonial en vivenda (1a/2a residència)", "POE Patrimoni", "9DC3E6", "00B0F0"),
            ReglaCorreccio("patrimoni_pista_poe", "MITJANA", "Delicte en pista → POE Pista", "POE Pista",
                "Delicte en autopista/autovia principal", "POE Pista", "B4C7E7", "2E5C8A"),
            ReglaCorreccio("centre_educatiu_poe", "MITJANA", "Delicte en educatiu → POE Educatiu", "POE Centres educatius",
                "Delicte en centre educatiu o als seus voltants", "POE Centres", "F4B084", "D97634"),
            ReglaCorreccio("habitatge_exclos_sense_ambit", "MITJANA", "Habitatge exclos → Sense àmbit", "Sense àmbit",
                "Habitatge exclos de POE (no residencial)", "Habitatge", "E8E8E8", "A6A6A6"),
        ]

    def _hauria_excloure_registre(self, registre: Dict) -> bool:
        motiu = str(registre.get("Motiu resolució", "")).lower()
        return any(excl in motiu for excl in self.motius_exclusio)

    def _esta_protegit(self, ambit: str) -> bool:
        return bool(ambit and ambit in self.ambits_protegits)

    def _es_delicte_patrimonial(self, tipus_fet: str) -> bool:
        return bool(tipus_fet and any(i in str(tipus_fet).lower() for i in ["robatori", "furt", "danys"]))

    def _es_delicte_poe_pista(self, tipus_fet: str) -> bool:
        return bool(tipus_fet and any(i in str(tipus_fet).lower() for i in ["furt", "robatori amb força", "teloners", "peruans"]))

    def _es_via_poe_pista(self, nom_vial: str, tipus_vial: str) -> bool:
        if not nom_vial and not tipus_vial:
            return False
        text = f"{tipus_vial} {nom_vial}".lower()
        for excl in ["tren", "metro", "renfe", "transport", "ferrocarril", "adif"]:
            if excl in text:
                return False
        return any(via in text for via in self.vies_poe_pista)

    def _es_punt_comercial(self, tipus_punt: str) -> bool:
        if not tipus_punt:
            return False
        tp = str(tipus_punt).lower()
        if tp.startswith("hab.") or "metro" in tp or "transport" in tp or "renfe" in tp:
            return False
        return any(i in tp for i in ["supermercat", "botiga", "restaurant", "benzinera", "mercat", "comerç", "comerc"])

    def _es_punt_habitatge(self, tipus_punt: str) -> bool:
        if not tipus_punt:
            return False
        tp = str(tipus_punt).strip().lower()
        if tp in self.tipus_punt_habitatge_exclos:
            return False
        if tp in self.tipus_punt_habitatge_valids:
            return True
        if ("1a res." in tp or "2a res." in tp) and not any(p in tp for p in ["pàrquing", "traster", "escala", "portal", "ascensor", "jardí", "pati"]):
            return True
        return False

    def _es_habitatge_exclos(self, tipus_punt: str) -> bool:
        return bool(tipus_punt and str(tipus_punt).strip().lower() in self.tipus_punt_habitatge_exclos)

    def _es_centre_educatiu(self, tipus_punt: str) -> bool:
        return bool(tipus_punt and any(i in str(tipus_punt).lower() for i in ["escola", "institut", "universitat"]))

    def _es_punt_oci_nocturn(self, tipus_punt: str) -> bool:
        return bool(tipus_punt and any(i in str(tipus_punt).lower() for i in ["discoteca", "pub", "club", "sala de festes"]))

    def _es_horari_nocturn(self, hora_inici: str) -> bool:
        if pd.isna(hora_inici) or not str(hora_inici).strip():
            return False
        h_ini = _get_config("horari_nocturn.hora_inici", 22)
        h_fi = _get_config("horari_nocturn.hora_fi", 7)
        try:
            dt = pd.to_datetime(str(hora_inici), errors="coerce")
            if pd.notna(dt):
                return dt.hour >= h_ini or dt.hour < h_fi
        except Exception:
            pass
        try:
            m = self.HORA_REGEX.search(str(hora_inici).strip().lower())
            if m:
                hora = int(m.group(1))
                return hora >= h_ini or hora < h_fi
        except (ValueError, TypeError, AttributeError):
            pass
        return False

    def _obte_camp_registre(self, registre: Dict, camp: str, variants: Optional[List[str]] = None) -> str:
        camps = [camp] + (variants or [])
        for c in camps:
            val = registre.get(c, "")
            if val is not None and str(val).strip():
                return str(val).strip().lower()
        return ""

    def aplicar_regles(self, registre: Dict) -> Optional[Dict]:
        try:
            if self._hauria_excloure_registre(registre):
                return None
            tipus_fet = str(registre.get("Tipus fet (nivell 2)", ""))
            tipus_punt_raw = registre.get("Tipus punt", "")
            tipus_punt = str(tipus_punt_raw).strip().lower()
            ambit_original = str(registre.get("Ambit procediment", "")).strip()
            hora_inici = registre.get("Hora (franja) Inici (fet)", "")
            nom_vial = self._obte_camp_registre(registre, "Nom Vial del fet", ["Nom Vial  del fet", "Nom Vial del fet"])
            tipus_vial = str(registre.get("Tipus vial", "")).lower()

            if self._esta_protegit(ambit_original):
                return None

            if ambit_original in ["Sense àmbit", "INTERNET"]:
                mapa = {"estafa_informatica": "estafa_informatica_estafes", "estafa_targetes": "estafa_targetes_estafes",
                        "estafa_ciberengany": "estafa_ciberengany_estafes", "estafa_basica": "estafa_basica_estafes",
                        "estafa_generica": "estafa_generica_estafes"}
                tipus_estafa = self._detecta_tipus_estafa(tipus_fet)
                if tipus_estafa and tipus_estafa in mapa:
                    return self._crear_resultat_correccio(mapa[tipus_estafa], registre)

            if ambit_original == "POE Habitatge" and self._es_habitatge_exclos(tipus_punt):
                return self._crear_resultat_correccio("habitatge_exclos_sense_ambit", registre)

            if self._es_delicte_patrimonial(tipus_fet) and self._es_punt_oci_nocturn(tipus_punt) and self._es_horari_nocturn(hora_inici):
                if ambit_original != "POE Oci nocturn":
                    return self._crear_resultat_correccio("oci_nocturn_poe", registre)

            if ambit_original == "Sense àmbit":
                if self._es_delicte_patrimonial(tipus_fet):
                    if self._es_delicte_poe_pista(tipus_fet) and self._es_via_poe_pista(nom_vial, tipus_vial):
                        return self._crear_resultat_correccio("patrimoni_pista_poe", registre)
                    if self._es_punt_oci_nocturn(tipus_punt) and self._es_horari_nocturn(hora_inici):
                        return self._crear_resultat_correccio("oci_nocturn_poe", registre)
                    if self._es_punt_comercial(tipus_punt):
                        return self._crear_resultat_correccio("patrimoni_comerc_poe", registre)
                    if self._es_punt_habitatge(tipus_punt):
                        return self._crear_resultat_correccio("patrimoni_habitatge_poe", registre)
                if self._es_centre_educatiu(tipus_punt):
                    return self._crear_resultat_correccio("centre_educatiu_poe", registre)

            return None
        except Exception as e:
            if self._log_cb:
                self._log_cb(f"Error aplicant regles: {e}\n")
            return None

    def _crear_resultat_correccio(self, id_regla: str, registre: Dict) -> Optional[Dict]:
        if id_regla in self.regles_desactivades:
            return None
        regla = next((r for r in self.regles if r.id_regla == id_regla), None)
        if not regla:
            return None
        ambit_original = str(registre.get("Ambit procediment", "")).strip()
        return {"id_regla": id_regla, "prioritat": regla.prioritat, "ambit_original": ambit_original,
                "ambit_nou": regla.ambit_desti, "motiu": regla.motiu, "categoria": regla.categoria,
                "color_original": regla.color_original, "color_suggerit": regla.color_suggerit}


# ============================================================================
# PROCESSAMENT  (adaptat a bytes en memòria, sense fitxers)
# ============================================================================


def processa_bytes(data: bytes, log_cb=None, progress_cb=None) -> Tuple[pd.DataFrame, pd.DataFrame, Dict]:
    """Processa un Excel rebut com a bytes. Retorna (df, df_correccions, stats)."""
    skiprows = _get_config("excel.skiprows", 5)
    validar = _get_config("excel.validar_columnes", True)
    unitats_filtrar = _obte_unitats_filtrar()

    def _log(msg: str) -> None:
        if log_cb:
            log_cb(msg)

    _log("\nProcessant arxiu...\n")

    try:
        df = pd.read_excel(BytesIO(data), skiprows=skiprows)
    except pd.errors.EmptyDataError:
        raise ValueError("L'arxiu Excel està buit")
    except pd.errors.ParserError as e:
        raise ValueError(f"Format Excel invàlid: {e}")

    df = _aplica_alias_columnes(df)

    mapeo_cols = _get_config("mapeo_columnes", {}) or {}
    rename_cols = {k: v for k, v in mapeo_cols.items() if k and v and k != v and k in df.columns}
    if rename_cols:
        df = df.rename(columns=rename_cols)
        _log(f"Columnes remapejades: {rename_cols}\n")

    if unitats_filtrar and COL_UNITAT in df.columns:
        df = df[df[COL_UNITAT].astype(str).str.strip().isin(unitats_filtrar)].copy()
        _log(f"Filtrat a {len(df)} registres ({len(unitats_filtrar)} unitats des de config)\n")
    else:
        _log(f"Carregats {len(df)} registres\n")

    if validar:
        ok, faltants = _valida_columnes(df)
        if not ok:
            raise ValueError(f"L'Excel no conté les columnes esperades. Faltants: {', '.join(faltants)}")
    _avisa_columnes_recomanades(df, log_cb)

    desactivades = _get_config("regles_desactivades", []) or []
    corrector = CorregidorAmbits(log_cb=log_cb, regles_desactivades=desactivades)
    if desactivades:
        _log(f"Regles desactivades per l'usuari: {len(desactivades)}\n")
    correccions = []
    df["Ambit corregit"] = df["Ambit procediment"]

    total = len(df)
    _log("Aplicant regles...\n")
    for i, (idx, registre) in enumerate(df.iterrows(), 1):
        if progress_cb and (i % 25 == 0 or i == total):
            progress_cb(i, total)
        resultat = corrector.aplicar_regles(registre.to_dict())
        if resultat:
            df.at[idx, "Ambit corregit"] = resultat["ambit_nou"]
            try:
                procediment_display = _format_procediment(registre.to_dict())
            except Exception:
                procediment_display = str(registre.get("Procediment", ""))

            correccions.append({
                "Comissaria": registre.get(COL_UNITAT, ""),
                "Procediment": procediment_display,
                "Data": registre.get("Data Inici (fet)", ""),
                "Àmbit Original": resultat["ambit_original"],
                "Motiu": resultat["motiu"],
                "Àmbit Suggerit": resultat["ambit_nou"],
                "Tipus Fet": registre.get("Tipus fet (nivell 2)", ""),
                "Tipus Punt": registre.get("Tipus punt", ""),
            })

    df_corr = pd.DataFrame(correccions)
    if len(df_corr) > 0:
        df_corr["Data_sort"] = pd.to_datetime(df_corr["Data"], errors="coerce")
        df_corr = df_corr.sort_values(["Comissaria", "Data_sort"], ascending=[True, False]).drop("Data_sort", axis=1).reset_index(drop=True)

    stats = {
        "total_registres": len(df),
        "correccions_totals": len(correccions),
        "pct_corregits": round((len(correccions) / len(df) * 100), 1) if len(df) > 0 else 0,
        "data_proces": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }
    return df, df_corr, stats


def _valor_excel(val: Any) -> Any:
    if pd.isna(val):
        return None
    if hasattr(val, "to_pydatetime"):
        try:
            return val.to_pydatetime()
        except Exception:
            return val
    return val


def genera_sortida_bytes(df_original: pd.DataFrame, df_correccions: pd.DataFrame, stats: Dict,
                         logo_path: Optional[str] = None, log_cb=None) -> bytes:
    """Genera l'Excel de sortida i el retorna com a bytes (per descarregar al navegador).

    logo_path: ruta dins del sistema de fitxers virtual de Pyodide (opcional).
    """
    def _log(msg: str) -> None:
        if log_cb:
            log_cb(msg)

    cfg = CONFIG.get("informe_excel", {})
    _log("Generant fitxer de sortida...\n")

    wb = Workbook()
    ws_orig = wb.active
    ws_orig.title = cfg.get("nom_full_original", "Dades Originals")
    for col_idx, col_name in enumerate(df_original.columns, 1):
        ws_orig.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row in enumerate(df_original.values, 2):
        for col_idx, val in enumerate(row, 1):
            ws_orig.cell(row=row_idx, column=col_idx, value=_valor_excel(val))

    ws_inf = wb.create_sheet(cfg.get("nom_full_informe", "Informe"))
    ws_inf.merge_cells(cfg.get("celles_logo", "A1:G1"))
    ws_inf.row_dimensions[1].height = cfg.get("fila_logo_altura", 60)

    if logo_path:
        try:
            import os
            if os.path.exists(logo_path) and os.path.getsize(logo_path) > 0:
                img = XLImage(logo_path)
                img.width = cfg.get("logo_ample", 400)
                img.height = cfg.get("logo_altura", 60)
                ws_inf.add_image(img, "B1")
                _log("Logo inserit\n")
        except Exception as e:
            _log(f"Error logo: {e}\n")

    ws_inf.merge_cells("B2:I2")
    t = ws_inf["B2"]
    t.value = "SUGGERÈNCIES CORRECCIONS ÀMBITS TRAMITACIÓ D'ATESTATS POLICIALS"
    t.font = Font(bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_inf.row_dimensions[2].height = cfg.get("fila_titol_altura", 25)

    ws_inf.merge_cells("B3:I3")
    st = ws_inf["B3"]
    st.value = f"Versió {VERSION} - Pauta Estafes Gener 2026"
    st.font = Font(bold=True, size=10, color="1F4E78", italic=True)
    st.alignment = Alignment(horizontal="center", vertical="center")
    ws_inf.row_dimensions[3].height = cfg.get("fila_subtitol_altura", 18)

    headers = ["Comissaria", "Procediment", "Data", "Àmbit Original", "Motiu", "Àmbit Suggerit", "Tipus Fet", "Tipus Punt"]
    hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col_idx, h in enumerate(headers, 2):
        c = ws_inf.cell(row=4, column=col_idx, value=h)
        c.fill, c.font = hf, Font(bold=True, color="FFFFFF", size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_inf.row_dimensions[4].height = cfg.get("fila_capcalera_altura", 20)

    anchos = cfg.get("ancho_columnes", {"B": 50, "C": 14, "D": 18, "E": 25, "F": 58, "G": 25, "H": 42, "I": 35})
    for letra, w in anchos.items():
        ws_inf.column_dimensions[letra].width = w

    if len(df_correccions) > 0:
        for row_idx, (_, fila) in enumerate(df_correccions.iterrows(), 5):
            ws_inf.cell(row=row_idx, column=2, value=_valor_excel(fila["Comissaria"]))
            ws_inf.cell(row=row_idx, column=3, value=_valor_excel(fila["Procediment"]))
            ws_inf.cell(row=row_idx, column=4, value=_valor_excel(fila["Data"]))
            ws_inf.cell(row=row_idx, column=5, value=_valor_excel(fila["Àmbit Original"]))
            ws_inf.cell(row=row_idx, column=6, value=_valor_excel(fila["Motiu"]))
            ws_inf.cell(row=row_idx, column=7, value=_valor_excel(fila["Àmbit Suggerit"]))
            ws_inf.cell(row=row_idx, column=8, value=_valor_excel(fila["Tipus Fet"]))
            ws_inf.cell(row=row_idx, column=9, value=_valor_excel(fila["Tipus Punt"]))

            fill_orig = PatternFill(start_color=get_color_by_ambit(fila["Àmbit Original"]), end_color=get_color_by_ambit(fila["Àmbit Original"]), fill_type="solid")
            fill_sugg = PatternFill(start_color=get_color_by_ambit(fila["Àmbit Suggerit"]), end_color=get_color_by_ambit(fila["Àmbit Suggerit"]), fill_type="solid")
            ws_inf.cell(row=row_idx, column=5).fill = fill_orig
            ws_inf.cell(row=row_idx, column=7).fill = fill_sugg

        ws_inf.auto_filter.ref = f"B4:I{4 + len(df_correccions)}"
    else:
        ws_inf.auto_filter.ref = "B4:I4"

    out = BytesIO()
    wb.save(out)
    _log("Fitxer generat correctament.\n")
    return out.getvalue()


def resum_text(stats: Dict, df_correccions: pd.DataFrame) -> str:
    """Retorna un resum en text del processament (per mostrar a la web)."""
    linies = [
        "=" * 40,
        "RESUM DEL PROCESSAMENT",
        "=" * 40,
        f"Total registres: {stats['total_registres']}",
        f"Correccions: {stats['correccions_totals']} ({stats['pct_corregits']}%)",
    ]
    if len(df_correccions) > 0:
        linies.append("")
        linies.append("Per comissaria:")
        for c, n in df_correccions.groupby("Comissaria").size().sort_values(ascending=False).items():
            linies.append(f"  - {c}: {n}")
    linies.append("=" * 40)
    return "\n".join(linies)


def correccions_to_records(df_correccions: pd.DataFrame) -> List[Dict]:
    """Converteix les correccions a una llista de dicts (per mostrar una taula a la web)."""
    if df_correccions is None or len(df_correccions) == 0:
        return []
    df = df_correccions.copy()
    if "Data" in df.columns:
        df["Data"] = df["Data"].apply(
            lambda v: pd.to_datetime(v).strftime("%d/%m/%Y") if pd.notna(pd.to_datetime(v, errors="coerce")) else ""
        )
    return df.fillna("").astype(str).to_dict("records")
